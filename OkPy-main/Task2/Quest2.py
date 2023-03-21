from openpyxl import load_workbook

# Load input workbook
input_workbook = load_workbook('input.xlsx')
input_worksheet = input_workbook.active

# Extract data and sort by name
data = [(row[0].strip(), row[3]) for row in input_worksheet.iter_rows(min_row=11, values_only=True)]
sorted_data = sorted(data, key=lambda x: x[0].split()[-1] if type(x[0]) == str else x[0])

# Write sorted data back to input worksheet
for i, (name, score) in enumerate(sorted_data):
    input_worksheet.cell(row=i+11, column=1, value=name)
    input_worksheet.cell(row=i+11, column=4, value=score)

# Save changes to input workbook
input_workbook.save('input.xlsx')