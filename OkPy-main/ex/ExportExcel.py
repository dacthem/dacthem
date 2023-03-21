import openpyxl as op

# khoi tao danh sach de luu TTSV
students = []

# dung for de nhap thong tin cho tung sinh vien va them TTSV vao danh sach
n = int(input("So luong sinh vien: "))
for i in range(n):
    masv = input("Nhap MaSV: ")
    hovaten = input("Nhap HoVaTen: ")
    diemtoan = float(input("Nhap DiemToan: "))
    diemly = float(input("Nhap DiemLy: "))
    diemhoa = float(input("Nhap DiemHoa: ")) 
    print("*")
    dtb = (diemtoan + diemly + diemhoa) / 3
    students.append([masv, hovaten, diemtoan, diemly, diemhoa, dtb])

# sap xep thong tin sinh vien theo DiemTB
students.sort(key=lambda x: x[5])

# khoi tao doi tuong workbook tu thu vien openpyxl
wb = op.Workbook()

# Tao sheet va them tieu de cho sheet
sheet = wb.active
sheet.title = "Danh sach sinh vien"
sheet.cell(row=1, column=1).value = "Ma Sinh Vien"
sheet.cell(row=1, column=2).value = "Ho va ten"
sheet.cell(row=1, column=3).value = "Diem Toan"
sheet.cell(row=1, column=4).value = "Diem Ly"
sheet.cell(row=1, column=5).value = "Diem Hoa"
sheet.cell(row=1, column=6).value = "DTB"

# Them thong tin sinh vien vao sheet
for i, student in enumerate(students):
    row = i + 2
    sheet.cell(row=row, column=1).value = student[0]
    sheet.cell(row=row, column=2).value = student[1]
    sheet.cell(row=row, column=3).value = student[2]
    sheet.cell(row=row, column=4).value = student[3]
    sheet.cell(row=row, column=5).value = student[4]
    sheet.cell(row=row, column=6).value = student[5]

# Lu thong tin sinh vien vao excel
wb.save("DanhSachSinhVien.xlsx")